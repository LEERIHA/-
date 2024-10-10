from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
import time
import pyautogui
import pyperclip
import PySimpleGUI as sg

# 엑셀 파일 불러오기
excel_file_path = "C:\Project\평가척도설정\excel.xlsx"
wb = load_workbook(excel_file_path)
sheet = wb.active  # 또는 원하는 시트를 지정

# 데이터를 저장할 리스트 초기화
evaluation_names = []
evaluation_scores = []
evaluation_descriptions = []

# 엑셀 데이터를 리스트에 저장
for row in sheet.iter_rows(min_row=1, values_only=True):  # 첫 번째 행은 헤더이므로 2부터 시작
    evaluation_names.append(row[0])  # A열의 데이터
    evaluation_scores.append(row[1])  # B열의 데이터
    evaluation_descriptions.append(row[2])  # C열의 데이터

# 리스트 내용 확인
print("평가명칭 리스트:", evaluation_names)
print("평가배점 리스트:", evaluation_scores)
print("평가척도설명 리스트:", evaluation_descriptions)

# GUI 레이아웃 생성
layout = [
    [sg.Text("사이트 주소"), sg.InputText(key="_URL_")],
    [sg.Text("아이디"), sg.InputText(key="_ID_")],
    [sg.Text("비밀번호"), sg.InputText(key="_PW_", password_char="*")],
    [sg.Text("평가척도제목"), sg.InputText(key="_TEXT1_")],
    [sg.Text("평가척도설명"), sg.InputText(key="_TEXT2_")],
    [sg.Submit(), sg.Cancel()],
]

# GUI 생성
window = sg.Window("자동화 프로그램", layout)

# 사용자로부터 입력 받기
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == "Cancel":
        break
    elif event == "Submit":
        url = values["_URL_"]
        user_id = values["_ID_"]
        user_pw = values["_PW_"]
        list_title = values["_TEXT1_"]
        list_explanation = values["_TEXT2_"]
        break


# 팝업창을 띄우기 위한 함수
def show_popup(message):
    sg.popup(message, title="알림")


# 저장할 메시지 리스트
message_list = []

# Chrome WebDriver 옵션 설정
chrome_options = Options()
chrome_options.add_argument("--ignore-certificate-errors")
chrome_options.add_argument("--ignore-ssl-errors")

# 웹페이지 열기
driver = webdriver.Chrome(options=chrome_options)
driver.implicitly_wait(5)
driver.maximize_window()

# 브라우저 실행
driver.get(url)

# 아이디 입력
id = driver.find_element(By.CSS_SELECTOR, "#id")  # 아이디 입력창
id.click()
pyperclip.copy(user_id)
id.click()
pyautogui.hotkey("ctrl", "v")
time.sleep(2)

# 비밀번호 입력
pw = driver.find_element(By.CSS_SELECTOR, "#password")
pw.click()
pyperclip.copy(user_pw)
pw.click()
pyautogui.hotkey("ctrl", "v")
time.sleep(2)

# 로그인 버튼
login_btn = driver.find_element(By.CSS_SELECTOR, "#btn-login")
login_btn.click()
time.sleep(5)

# 이용서비스 선택 화면 닫기 클릭
first_btn = driver.find_element(
    By.CSS_SELECTOR, "#modal > div > div.caching-set > span > span > span > button"
)
first_btn.click()

# 관리자 사이트 페이지로 이동
admin_btn = driver.find_element(
    By.CSS_SELECTOR,
    "#wrap > main > div > section > a:nth-child(2) > div > div > div.service-footer.float-menu",
)
admin_btn.click()

# 채용 전형 관리 클릭
model_btn = driver.find_element(
    By.CSS_SELECTOR, "#sidebarFaceLift > ul > li:nth-child(3) > a"
)
model_btn.click()
time.sleep(2)

# 평가척도설정 페이지 이동
model_btn = driver.find_element(
    By.CSS_SELECTOR, "#sidebarFaceLift > ul > li.open > div > ul > li:nth-child(3) > a"
)
model_btn.click()
time.sleep(2)

# 평가척도 추가 버튼
useList_btn = driver.find_element(
    By.CSS_SELECTOR, "#useList > div > div > button:nth-child(2)"
)
useList_btn.click()
time.sleep(2)

# 평가척도 관리 제목 입력
list_title_btn = driver.find_element(
    By.CSS_SELECTOR,
    "#modalBody > table:nth-child(2) > tbody > tr:nth-child(1) > td > input",
)
list_title_btn.click()
pyperclip.copy(list_title)
pyautogui.hotkey("ctrl", "v")
time.sleep(2)


# 평가척도 관리 > 설명 입력
list_explanation_btn = driver.find_element(
    By.CSS_SELECTOR,
    "#modalBody > table:nth-child(2) > tbody > tr:nth-child(2) > td > textarea",
)
list_explanation_btn.click()
pyperclip.copy(list_explanation)
pyautogui.hotkey("ctrl", "v")
time.sleep(2)

# 평가척도 데이터를 반복하여 입력 및 저장
for i in range(len(evaluation_names)):
    # 평가척도 명칭 입력
    title_input = driver.find_element(
        By.CSS_SELECTOR, "#modalTable > tbody > tr > td:nth-child(3) > input"
    )
    title_input.clear()
    title_input.send_keys(evaluation_names[i])

    # 평가배점 입력
    score_input = driver.find_element(
        By.CSS_SELECTOR, "#modalTable > tbody > tr > td:nth-child(4) > input"
    )
    score_input.clear()
    score_input.send_keys(str(evaluation_scores[i]))

    # 평가척도 기준 설명 입력
    description_input = driver.find_element(
        By.CSS_SELECTOR, "#modalTable > tbody > tr > td:nth-child(5) > input"
    )
    description_input.clear()
    description_input.send_keys(evaluation_descriptions[i])
    time.sleep(2)

    # 수정버튼 클릭
    change_btn = driver.find_element(
        By.CSS_SELECTOR,
        "#modalTable > tbody > tr:nth-child(i) > td:nth-child(7) > button",
    )
    change_btn.click()
    time.sleep(2)

    # 확인버튼 클릭
    ok_btn = driver.find_element(By.CSS_SELECTOR, "#Dialog > div > button:nth-child(1)")
    ok_btn.click()
    time.sleep(2)


# 저장버튼 클릭
save_btn = driver.find_element(By.CSS_SELECTOR, "#modalSubmit")
save_btn.click()
time.sleep(2)
